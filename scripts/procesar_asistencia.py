#!/usr/bin/env python3
"""
Procesa archivos XLSX de asistencia/disponibilidad de Boletomóvil
y genera data/asistencia.json para el dashboard de Series Inaugurales LMB.

Uso:
    python scripts/procesar_asistencia.py [carpeta_xlsx]

Si no se especifica carpeta, usa data/asistencia/ por defecto.
"""

import json
import os
import sys
import re

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl


def safe_int(v):
    """Convierte valor a int de forma segura."""
    if v is None or str(v).strip() in ('', '-', 'None'):
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def procesar_asistencia_xlsx(filepath):
    """Procesa un archivo XLSX de asistencia/disponibilidad."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # Evento (fila 3)
    evento = ''
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        evento = str(row[0]).strip() if row[0] else ''
        break

    # Headers (fila 5)
    headers = []
    for row in ws.iter_rows(min_row=5, max_row=5, values_only=True):
        headers = [str(h).strip() if h else f'col{i}' for i, h in enumerate(row)]
        break

    # Mapear columnas
    col = {}
    for i, h in enumerate(headers):
        col[h] = i

    # Procesar zonas (fila 6+)
    zonas = []
    aforo_total = 0
    ocupados_total = 0
    disponibles_total = 0
    bloqueos_total = 0
    epv_total = 0

    for row in ws.iter_rows(min_row=6, values_only=True):
        zona_name = str(row[0]).strip() if row[0] else ''
        if not zona_name or zona_name in ('', 'TOTALES'):
            continue

        def get(name, default=0):
            idx = col.get(name)
            if idx is not None and idx < len(row):
                return safe_int(row[idx])
            return default

        # Handle different column names between years
        epv = get('EN PROCESO DE VENTA') or get('EN PROCESO')
        vendidos = get('VENDIDOS CON PRECIO') or get('VENDIDOS')
        promos = get('PROMOS')
        cortesias = get('CORTESÍAS')
        precio_1 = get('PRECIO $1')
        abonos_vend = get('ABONOS VENDIDOS')
        abonos_prom = get('ABONOS PROMOS')
        abonos_cort = get('ABONOS CORTESÍAS')
        abonos_p1 = get('ABONOS PRECIO $1')
        paquetes = get('PAQUETES')
        paq_cort = get('PAQUETES CORTESIAS')
        paq_p1 = get('PAQUETES PRECIO $1')
        promociones = get('PROMOCIONES')
        bloqueos = get('BLOQUEOS')
        aforo = get('AFORO')

        # Use TOTAL/OCUPADOS column if available, otherwise calculate
        ocupados_directo = get('TOTAL') or get('OCUPADOS')
        ocupados_calculado = (vendidos + promos + cortesias + precio_1 +
                    abonos_vend + abonos_prom + abonos_cort + abonos_p1 +
                    paquetes + paq_cort + paq_p1 + promociones)
        ocupados = ocupados_directo if ocupados_directo > 0 else ocupados_calculado

        # Use DISPONIBLES column if available, otherwise calculate
        disp_directo = get('DISPONIBLES')
        disponibles = disp_directo if disp_directo > 0 else max(0, aforo - ocupados - bloqueos - epv)
        pct = round(ocupados / aforo * 100, 1) if aforo > 0 else 0

        zonas.append({
            'zona': zona_name,
            'aforo': aforo,
            'vendidos_precio': vendidos,
            'promos': promos,
            'cortesias': cortesias,
            'precio_1': precio_1,
            'abonos': abonos_vend + abonos_prom + abonos_cort + abonos_p1,
            'paquetes': paquetes + paq_cort + paq_p1,
            'promociones': promociones,
            'bloqueos': bloqueos,
            'en_proceso': epv,
            'ocupados': ocupados,
            'disponibles': disponibles,
            'pct_ocupacion': pct
        })

        aforo_total += aforo
        ocupados_total += ocupados
        disponibles_total += disponibles
        bloqueos_total += bloqueos
        epv_total += epv

    wb.close()

    pct_total = round(ocupados_total / aforo_total * 100, 1) if aforo_total > 0 else 0

    return {
        'evento': evento,
        'aforo_total': aforo_total,
        'ocupados_total': ocupados_total,
        'disponibles_total': disponibles_total,
        'bloqueos_total': bloqueos_total,
        'en_proceso_total': epv_total,
        'pct_ocupacion': pct_total,
        'zonas': zonas
    }


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_dir = os.path.dirname(script_dir)

    if len(sys.argv) > 1:
        asist_dir = sys.argv[1]
    else:
        asist_dir = os.path.join(repo_dir, 'data', 'asistencia')

    if not os.path.isdir(asist_dir):
        print(f"Carpeta de asistencia no encontrada: {asist_dir}")
        print("Saltando procesamiento de asistencia.")
        return

    # Buscar archivos XLSX
    archivos = [f for f in os.listdir(asist_dir) if f.endswith('.xlsx')]

    if not archivos:
        print(f"No se encontraron archivos .xlsx en {asist_dir}")
        return

    print(f"Procesando {len(archivos)} archivos de asistencia de {asist_dir}...")

    resultado = {}
    for archivo in sorted(archivos):
        # Formato: "Charros 2025.xlsx" o "Rieleros 2026.xlsx"
        match = re.match(r'(.+?)\s+(\d{4})\.xlsx', archivo)
        if not match:
            print(f"  Saltando {archivo} (formato no reconocido)")
            continue

        equipo = match.group(1).strip()
        year = match.group(2)
        filepath = os.path.join(asist_dir, archivo)

        print(f"  {equipo} {year}...", end=' ')
        try:
            datos = procesar_asistencia_xlsx(filepath)
            if equipo not in resultado:
                resultado[equipo] = {}
            resultado[equipo][year] = datos
            print(f"OK ({datos['aforo_total']} aforo, {datos['ocupados_total']} ocupados, {datos['pct_ocupacion']}%)")
        except Exception as e:
            print(f"ERROR: {e}")

    # Guardar JSON
    output_path = os.path.join(repo_dir, 'data', 'asistencia.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f"\nGenerado: {output_path}")
    print(f"Equipos: {len(resultado)}")
    for equipo, years in sorted(resultado.items()):
        for year, data in sorted(years.items()):
            print(f"  {equipo} {year}: {data['pct_ocupacion']}% ocupación ({data['ocupados_total']}/{data['aforo_total']})")


if __name__ == '__main__':
    main()
