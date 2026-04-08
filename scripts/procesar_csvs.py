#!/usr/bin/env python3
"""
Procesa archivos XLSX de Boletomóvil y genera data/2026.json
para el dashboard de Series Inaugurales LMB.

Uso:
    python scripts/procesar_csvs.py [carpeta_xlsx]

Si no se especifica carpeta, usa data/csv/ por defecto.
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl


def es_evento_valido(evento):
    """Filtra eventos reales vs basura (vendedores, admin, etc.)"""
    if not evento or not isinstance(evento, str):
        return False
    evento_lower = evento.strip().lower()
    # Eventos basura conocidos
    basura = ['-', 'administradores', 'venta en línea', 'acceso qr bm']
    if evento_lower in basura:
        return False
    # Si empieza con [deshabilitado: es basura
    if evento_lower.startswith('[deshabilitado'):
        return False
    # Eventos válidos contienen estas palabras clave
    keywords = ['inaugural', 'paquete', 'serie', 'juego', 'j1:', 'j2:', 'j3:']
    if any(kw in evento_lower for kw in keywords):
        return True
    # Si no matchea keywords pero tiene muchas palabras con "vs", probablemente es válido
    if ' vs ' in evento_lower:
        return True
    return False


def parsear_fecha(fecha_str):
    """Convierte DD/MM/YY HH:MM → date string YYYY-MM-DD"""
    if not fecha_str or not isinstance(fecha_str, str):
        return None
    try:
        # Formato: 06/04/26 11:12
        dt = datetime.strptime(fecha_str.strip(), '%d/%m/%y %H:%M')
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        try:
            # Intentar sin hora
            dt = datetime.strptime(fecha_str.strip()[:8], '%d/%m/%y')
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            return None


def parsear_subtotal(valor):
    """Convierte subtotal a float. '-' = 0"""
    if valor is None or valor == '-' or valor == '':
        return 0.0
    try:
        return float(valor)
    except (ValueError, TypeError):
        return 0.0


def procesar_xlsx(filepath):
    """Procesa un archivo XLSX de Boletomóvil y retorna los datos del equipo."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # Leer headers (fila 3)
    headers = []
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        headers = [str(h).strip() if h else '' for h in row]
        break

    # Mapear índices de columnas
    col_map = {}
    for i, h in enumerate(headers):
        col_map[h] = i

    idx_orden = col_map.get('NÚMERO DE ORDEN', 0)
    idx_evento = col_map.get('EVENTO', 2)
    idx_fecha = col_map.get('FECHA', 3)
    idx_vendido_por = col_map.get('VENDIDO POR', None)
    idx_tipo = col_map.get('TIPO', None)
    idx_medio = col_map.get('MEDIO DE COMPRA', 9)
    idx_subtotal = col_map.get('SUBTOTAL', 13)
    idx_zona = col_map.get('ZONA', None)

    # Procesar filas
    eventos_set = set()
    ordenes_todas = set()
    ordenes_online = set()
    ordenes_taquilla = set()
    total_subtotal = 0.0
    total_online = 0.0
    total_taquilla = 0.0
    total_paquete = 0.0
    total_no_paquete = 0.0
    boletos = 0
    boletos_online = 0
    boletos_taquilla = 0
    boletos_cortesias = 0
    boletos_cortesias_online = 0
    boletos_cortesias_taquilla = 0
    cortesias_por_persona = defaultdict(lambda: {'total': 0, 'online': 0, 'taquilla': 0})
    daily_data = defaultdict(lambda: {'subtotal': 0.0, 'boletos': 0})
    zona_data = defaultdict(lambda: {
        'boletos': 0, 'subtotal': 0.0,
        'boletos_online': 0, 'subtotal_online': 0.0,
        'boletos_taquilla': 0, 'subtotal_taquilla': 0.0,
        'boletos_cortesias': 0, 'boletos_cortesias_online': 0, 'boletos_cortesias_taquilla': 0
    })
    fechas = []
    has_paquete = False

    for row in ws.iter_rows(min_row=4, values_only=True):
        orden = row[idx_orden] if len(row) > idx_orden else None
        evento = row[idx_evento] if len(row) > idx_evento else None
        fecha_str = row[idx_fecha] if len(row) > idx_fecha else None
        vendido_por_raw = row[idx_vendido_por] if idx_vendido_por is not None and len(row) > idx_vendido_por else None
        tipo_raw = row[idx_tipo] if idx_tipo is not None and len(row) > idx_tipo else None
        medio = row[idx_medio] if len(row) > idx_medio else None
        subtotal_raw = row[idx_subtotal] if len(row) > idx_subtotal else None
        zona_raw = row[idx_zona] if idx_zona is not None and len(row) > idx_zona else None

        # Filtrar filas sin orden o con evento basura
        if orden is None:
            continue
        if not es_evento_valido(str(evento) if evento else ''):
            continue

        eventos_set.add(str(evento))
        subtotal = parsear_subtotal(subtotal_raw)
        fecha = parsear_fecha(str(fecha_str) if fecha_str else '')
        zona = str(zona_raw).strip() if zona_raw and str(zona_raw).strip() not in ('-', '', 'None') else None

        # Detectar cortesía: por columna TIPO o por subtotal = 0
        tipo_str = str(tipo_raw).strip().lower() if tipo_raw else ''
        es_cortesia = tipo_str == 'cortesía' or tipo_str == 'cortesia' or (subtotal == 0 and tipo_str not in ('adulto',))

        # Clasificar online vs taquilla
        es_taquilla = (str(medio).strip().lower() == 'taquilla') if medio else False
        es_paquete_flag = 'paquete' in str(evento).lower()
        if es_paquete_flag:
            has_paquete = True

        # Acumular
        boletos += 1
        total_subtotal += subtotal
        ordenes_todas.add(orden)

        if es_cortesia:
            boletos_cortesias += 1
            if es_taquilla:
                boletos_cortesias_taquilla += 1
            else:
                boletos_cortesias_online += 1
            # Track by person
            vendido_por = str(vendido_por_raw).strip() if vendido_por_raw and str(vendido_por_raw).strip() not in ('-', '', 'None') else 'Sin asignar'
            cortesias_por_persona[vendido_por]['total'] += 1
            if es_taquilla:
                cortesias_por_persona[vendido_por]['taquilla'] += 1
            else:
                cortesias_por_persona[vendido_por]['online'] += 1

        if es_taquilla:
            boletos_taquilla += 1
            total_taquilla += subtotal
            ordenes_taquilla.add(orden)
        else:
            boletos_online += 1
            total_online += subtotal
            ordenes_online.add(orden)

        if es_paquete_flag:
            total_paquete += subtotal
        else:
            total_no_paquete += subtotal

        if fecha:
            fechas.append(fecha)
            daily_data[fecha]['subtotal'] += subtotal
            daily_data[fecha]['boletos'] += 1

        # Acumular por zona
        if zona:
            zona_data[zona]['boletos'] += 1
            zona_data[zona]['subtotal'] += subtotal
            if es_cortesia:
                zona_data[zona]['boletos_cortesias'] += 1
                if es_taquilla:
                    zona_data[zona]['boletos_cortesias_taquilla'] += 1
                else:
                    zona_data[zona]['boletos_cortesias_online'] += 1
            if es_taquilla:
                zona_data[zona]['boletos_taquilla'] += 1
                zona_data[zona]['subtotal_taquilla'] += subtotal
            else:
                zona_data[zona]['boletos_online'] += 1
                zona_data[zona]['subtotal_online'] += subtotal

    wb.close()

    # Construir daily array ordenado
    daily = []
    for date_str in sorted(daily_data.keys()):
        daily.append({
            'date': date_str,
            'subtotal': round(daily_data[date_str]['subtotal'], 2),
            'boletos': daily_data[date_str]['boletos']
        })

    # Calcular fechas
    fechas_sorted = sorted(set(fechas))
    if fechas_sorted:
        first_dt = datetime.strptime(fechas_sorted[0], '%Y-%m-%d')
        last_dt = datetime.strptime(fechas_sorted[-1], '%Y-%m-%d')
        first_date = first_dt.strftime('%d/%m/%Y')
        last_date = last_dt.strftime('%d/%m/%Y')
        duration_days = (last_dt - first_dt).days
    else:
        first_date = ''
        last_date = ''
        duration_days = 0

    return {
        'subtotal': round(total_subtotal, 2),
        'subtotal_online': round(total_online, 2),
        'subtotal_taquilla': round(total_taquilla, 2),
        'subtotal_paquete': round(total_paquete, 2),
        'subtotal_no_paquete': round(total_no_paquete, 2),
        'boletos': boletos,
        'boletos_online': boletos_online,
        'boletos_taquilla': boletos_taquilla,
        'boletos_cortesias': boletos_cortesias,
        'boletos_cortesias_online': boletos_cortesias_online,
        'boletos_cortesias_taquilla': boletos_cortesias_taquilla,
        'cortesias_por_persona': sorted(
            [{'persona': p, 'total': d['total'], 'online': d['online'], 'taquilla': d['taquilla']}
             for p, d in cortesias_por_persona.items()],
            key=lambda x: x['total'], reverse=True
        ),
        'ordenes': len(ordenes_todas),
        'ordenes_online': len(ordenes_online),
        'ordenes_taquilla': len(ordenes_taquilla),
        'events': sorted(eventos_set),
        'has_paquete': has_paquete,
        'first_date': first_date,
        'last_date': last_date,
        'duration_days': duration_days,
        'daily': daily,
        'zonas': {z: {
            'boletos': d['boletos'],
            'boletos_pagados': d['boletos'] - d['boletos_cortesias'],
            'boletos_cortesias': d['boletos_cortesias'],
            'boletos_cortesias_online': d['boletos_cortesias_online'],
            'boletos_cortesias_taquilla': d['boletos_cortesias_taquilla'],
            'subtotal': round(d['subtotal'], 2),
            'ticket_promedio': round(d['subtotal'] / (d['boletos'] - d['boletos_cortesias']), 2) if (d['boletos'] - d['boletos_cortesias']) > 0 else 0,
            'boletos_online': d['boletos_online'],
            'subtotal_online': round(d['subtotal_online'], 2),
            'boletos_taquilla': d['boletos_taquilla'],
            'subtotal_taquilla': round(d['subtotal_taquilla'], 2)
        } for z, d in sorted(zona_data.items())}
    }


def main():
    # Determinar carpeta de entrada
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_dir = os.path.dirname(script_dir)

    if len(sys.argv) > 1:
        csv_dir = sys.argv[1]
    else:
        csv_dir = os.path.join(repo_dir, 'data', 'csv')

    if not os.path.isdir(csv_dir):
        print(f"Error: Carpeta no encontrada: {csv_dir}")
        sys.exit(1)

    # Procesar cada año disponible
    for year in ['2024', '2025', '2026']:
        archivos = [f for f in os.listdir(csv_dir) if f.endswith('.xlsx') and f' {year}.' in f]

        if not archivos:
            print(f"No se encontraron archivos *{year}*.xlsx en {csv_dir}")
            continue

        print(f"\nProcesando {len(archivos)} archivos de {year}...")

        resultado = {}
        for archivo in sorted(archivos):
            # Extraer nombre del equipo: "Bravos 2026.xlsx" → "Bravos"
            equipo = archivo.replace(f'_{year}.xlsx', '').replace(f' {year}.xlsx', '').strip()
            filepath = os.path.join(csv_dir, archivo)

            print(f"  {equipo}...", end=' ')
            try:
                datos = procesar_xlsx(filepath)
                resultado[equipo] = datos
                zonas_count = len(datos.get('zonas', {}))
                print(f"OK ({datos['boletos']} boletos, ${datos['subtotal']:,.0f}, {zonas_count} zonas)")
            except Exception as e:
                print(f"ERROR: {e}")

        # Guardar JSON
        output_path = os.path.join(repo_dir, 'data', f'{year}.json')
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(resultado, f, ensure_ascii=False, indent=2)

        print(f"Generado: {output_path}")
        print(f"Equipos: {len(resultado)}")
        total = sum(d['subtotal'] for d in resultado.values())
        print(f"Venta total {year}: ${total:,.0f}")


if __name__ == '__main__':
    main()
